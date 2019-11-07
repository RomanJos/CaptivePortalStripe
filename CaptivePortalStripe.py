from flask import Flask, render_template, request, jsonify, send_file
from flask_mail import Mail, Message
import time
import json
import sys
import os
import logging
import stripe  #The AUR package is not up to date
from openpyxl import load_workbook
from datetime import datetime
import codecs
import threading
import ast

#Get language
try:
	Language = sys.argv[1]
	if (Language == "French") or (Language == "English"):
		print (f"Language selected for backend : {Language}")
	else:
		raise ValueError()
except Exception:
	print ('Language available for backend are French and English.\nYou can set one with "python CaptivePortalStripe.py English" for example')
	sys.exit()

#Dictionary
with open("Dictionary.json") as f:
	Dictionary = json.load(f)
	Dictionary = Dictionary[Language]

#Flask app
app = Flask(__name__)

#Change logging to do our own
log = logging.getLogger('werkzeug')
log.setLevel(logging.ERROR)

#General Settings
with open("Settings.json") as f:
	Settings = json.load(f)

#Stripe settings
PublicKey = Settings["StripeKey"]["Public"]
PrivateKey = Settings["StripeKey"]["Private"]
stripe.api_key = PrivateKey
WeekCheckoutID = Settings["StripeCheckoutID"]["Week"]
MonthCheckoutID = Settings["StripeCheckoutID"]["Month"]
StripeSuccessPage = Settings["StripeSuccessPage"]
StripeCancelPage = Settings["StripeCancelPage"]

#Email settings
app.config.update(
	MAIL_SERVER=Settings["MailConfig"]["Server"],
	MAIL_PORT=int(Settings["MailConfig"]["Port"]),
	MAIL_USE_SSL=ast.literal_eval(Settings["MailConfig"]["SSL"]),#ast.literal_eval is to convert the string "True" to a bool True
	MAIL_USERNAME=Settings["MailConfig"]["UserName"],
	MAIL_PASSWORD=Settings["MailConfig"]["Password"])
mail = Mail(app)

#Create SessionID in case its not here
if not os.path.isfile("SessionID.json"):
	with open("SessionID.json", 'w') as f:
		f.write('{}')

#Load PaymentHistory
PaymentHistoryXlsx = "PaymentHistory.xlsx"
PaymentHistoryload = load_workbook(PaymentHistoryXlsx)
PaymentHistoryloadActive = PaymentHistoryload.active
#----------------------------------------------------------------------------------

#Return True if the string url is in the "Location"'settings list
def isThisInLocationTxt(url):
	for Location in Settings["Location"]:
		if Location == url:
			return True
	return False

#Return VoucherID and VoucherPassword based on the url and WeekOrMonth which can be "Week" or "Month"
def GetVoucher(url, WeekOrMonth):
	print(f"{Dictionary['Print']['SomeonePayFrom']} {url}")

	xlsx = f"Voucher/{url}-{WeekOrMonth}.xlsx"
	VoucherWorkBook = load_workbook(xlsx)
	VoucherList = VoucherWorkBook.active

	RowEmpty = rowvoucherpasused(VoucherList)

	VoucherID = str(returncellcontent("b", RowEmpty, VoucherList))
	VoucherPassword = str(returncellcontent("c", RowEmpty, VoucherList))

	writeincell("a", RowEmpty, "USED", VoucherList)
	VoucherWorkBook.save(xlsx)

	return VoucherID, VoucherPassword

#Send a mail to the boss
def ENVOYERMAILAdmin(objet, message):
	msg = Message(objet,
				  sender=Settings["MailConfig"]["SendMailAs"],
				  recipients=[Settings["MailConfig"]["Admin"]])
	msg.body = message
	mail.send(msg)

#Send a mail to clientmail
def ENVOYERMAILclient(clientmail, objet, message):
	print(f"{Dictionary['Print']['SendMailTo']} {clientmail}")
	msg = Message(objet,
				  sender=Settings["MailConfig"]["SendMailAs"],
				  recipients=[clientmail])
	msg.body = message
	mail.send(msg)

#Return the value of the cell in row and column inside VoucherList (for example to get VoucherID or VoucherPassword)
def returncellcontent(column, row, Voucherlist):
	value = Voucherlist[column + row].value
	return value

#Write contenttowrite in the cell in row and column inside Voucherlist
def writeincell(column, row, contenttowrite, Voucherlist):
	contenttowrite = str(contenttowrite)
	Voucherlist[column + row].value = contenttowrite

#Check if there is not "USED" in the first cell of a row and return this row's number
def rowvoucherpasused(Voucherlist):
	incre = 0
	while True:
		incre = incre + 1
		cell = returncellcontent("a", str(incre), Voucherlist)
		if "USED" in str(cell):
			continue
		return str(incre)#Need to be str

#Below is some function that I will probably remove in the future
#Return the value of the cell in row and column inside PaymentHistory (for example to get a empty row or just a cell value)
def returncellcontentPaymentHistory(column, row):
	global PaymentHistoryloadActive
	return PaymentHistoryloadActive[column + row].value

#Write contenttowrite in the cell in row and column inside PaymentHistory
def writeincellPaymentHistory(column, row, contenttowrite):
	global PaymentHistoryloadActive
	PaymentHistoryloadActive[column + row].value = contenttowrite

#Check if there is nothing in the second cell of a row and return this row's number
def getrowvidePaymentHistory():
	incre = 1  # 1 pour skip la première ligne
	while True:
		incre = incre + 1
		cell = returncellcontentPaymentHistory("b", str(incre))
		if str(cell) == 'None':
			return str(incre)

#Print something like this 13:4 8/10/2019 x.xx.x.xx https://url.com/
def printRequest(IP, Host, request):
	heure = datetime.now()
	print(f"{heure.hour}:{heure.minute} {heure.day}/{heure.month}/{heure.year} {IP} {Host}")

#------------------------------------------------------------------------


@app.route('/')
def root():
	Location = request.headers['HTTP_X_LOCATION']
	Host = request.headers['HTTP_X_HOST']
	IP = request.headers['HTTP_X_REAL_IP']
	printRequest(IP, Host, str(request))

	#Check before he buy or does anything else
	if not isThisInLocationTxt(Location):
		print(f"{Location} {Dictionary['Print']['NotInLocation']}")
		ENVOYERMAILAdmin(
			f"{Dictionary['Mail']['NotInLocationObject']} {Location}",
			f"{Dictionary['Mail']['NotInLocationBody'].replace('{IP}', IP).replace('{Host}', Host)}"
		)
		return f"{Dictionary['Flask']['ErrorMailSentHereIsFicelle']}"
	return render_template("Index.html", PKStripe=PublicKey, WeekCheckoutID=WeekCheckoutID, MonthCheckoutID=MonthCheckoutID, StripeSuccessPage=StripeSuccessPage, StripeCancelPage=StripeCancelPage)


@app.route('/Success')
def Success():
	Location = request.headers['HTTP_X_LOCATION']
	Host = request.headers['HTTP_X_HOST']  #This is not really needed
	IP = request.headers['HTTP_X_REAL_IP']
	printRequest(IP, Host, str(request))

	#Check if Its in Location.txt to be sure
	if not isThisInLocationTxt(Location):
		print(f"{Location} {Dictionary['Print']['NotInLocation']}")
		ENVOYERMAILAdmin(
			f"{Dictionary['Mail']['NotInLocationObject']} {Location}",
			f"{Dictionary['Mail']['NotInLocationBody'].replace('{IP}', IP).replace('{Host}', Host)}"
		)
		return f"{Dictionary['Flask']['ErrorMailSentHereIsFicelle']}"

	#basic check
	if not request.args.get('session_id'):
		return redirect("/")

	# here we want to get the value of session_id (i.e. ?session_id=some-value)
	session_id = request.args.get('session_id').replace("/", "")  #I don't know if I need to keep .replace("/", "")

	#Mofify ServerURL to include the Location
	ServerURL = Settings["ServerURL"].replace('{Location}', Location)

	#get the stripe things
	try:
		Result = stripe.checkout.Session.retrieve(session_id)
	except Exception as e:
		print(f"{Dictionary['Print']['WrongSession_ID']} {e}")
		ENVOYERMAILAdmin(
			f"{Dictionary['Mail']['WrongSession_IDObject']}",
			f"{Dictionary['Mail']['WrongSession_IDBody'].replace('{IP}', IP).replace('{Host}', Host).replace('{session_id}', session_id)}{e}"
		)
		return f"{Dictionary['Flask']['ErrorMailSentHereIsFicelle']}"

	#Here we check if its a subscription or a simple payment
	if Result.subscription == None:
		WeekOrMonth = 'Week'
	else:
		WeekOrMonth = 'Month'

	#prepare the json of session to avoid the client re-using the same session id
	with open("SessionID.json") as f:
		SessionIDJson = None
		SessionIDJson = json.load(f)
		if session_id in SessionIDJson:
			if (SessionIDJson[session_id]["Plan"] == "DELETED"):
				return render_template('Success.html',
									   VoucherID="DELETED",
									   VoucherPassword="DELETED",
									   info3=f"{Dictionary['Flask']['InfoPlanHasBeenDeleted']}"
				)
			if WeekOrMonth == "Week":
				return render_template(
					'Success.html',
					VoucherID=SessionIDJson[session_id]["VoucherID"],
					VoucherPassword=SessionIDJson[session_id]["VoucherPassword"],
					info=f"{Dictionary['Flask']['OneDeviceAtATime']}"
				)
			if WeekOrMonth == "Month":
				return render_template(
					'Success.html',
					VoucherID=SessionIDJson[session_id]["VoucherID"],
					VoucherPassword=SessionIDJson[session_id]["VoucherPassword"],
					info=f"{Dictionary['Flask']['OneDeviceAtATime']}",
					info2=f"{Dictionary['Flask']['IfYouWantToDeleteYourPlanGoInYourMail']}"
				)

	Customer = stripe.Customer.retrieve(Result.customer)
	Email = Customer.email
	print(Email)

	#Get voucher
	try:
		VoucherID, VoucherPassword = GetVoucher(Location, WeekOrMonth)
	except Exception as e:#If something goes wrong its better to have a trace
		print(f"{Dictionary['Print']['ExceptionGetVoucher'].replace('{e}', str(e))}")
		ENVOYERMAILAdmin(
			f"{Dictionary['Mail']['ExceptionGetVoucherObject']}",
			f"{Dictionary['Mail']['ExceptionGetVoucherBody'].replace('{Email}', Email).replace('{Host}', Host).replace('{e}', str(e))}"
		)
		return f"{Dictionary['Mail']['ErrorNoMoreVoucherMailSent']}"
	#If there is no more Voucher
	if (VoucherID == None) or (VoucherPassword == None):
		ENVOYERMAILAdmin(
			f"{Dictionary['Mail']['ErrorNoMoreVoucherObject']}",
			f"{Dictionary['Mail']['ErrorNoMoreVoucherBody'].replace('{Email}', Email).replace('{Host}', Host)}"
		)
		return f"{Dictionary['Mail']['ErrorNoMoreVoucherMailSent']}"
	print(f"{Dictionary['Print']['Voucher'].replace('{VoucherID}', VoucherID).replace('{VoucherPassword}', VoucherPassword)}")
	#Save the Stripe Session_id so if he refresh we don't provide him new voucher, I didn't find any other way to do it
	if SessionIDJson == None:
		SessionIDJson = {}
	if WeekOrMonth == "Month":
		SessionIDJson[session_id] = {
			"Location": Location,
			"Plan": "Month",
			"Email": Email,
			"sub_xxx": Result.subscription,
			"Customer": Customer.id,
			"VoucherID": VoucherID,
			"VoucherPassword": VoucherPassword
		}
	elif WeekOrMonth == "Week":
		SessionIDJson[session_id] = {
			"Location": Location,
			"Plan": "Week",
			"Email": Email,
			"payment_intent": Result.payment_intent,
			"Customer": Customer.id,
			"VoucherID": VoucherID,
			"VoucherPassword": VoucherPassword
		}
	SessionIDJson = json.dumps(SessionIDJson,
							   indent=4,
							   sort_keys=True,
							   ensure_ascii=False
							   )
	with open("SessionID.json", "w") as json_file:
		json_file.write(SessionIDJson)

	#Get the Admin mail before sending mail
	Admin = Settings["MailConfig"]["Admin"]

	#Send confirmation mail, if its month we send a link to delete his plan
	if WeekOrMonth == "Week":
		ENVOYERMAILAdmin(
			f"{Dictionary['Mail']['SucessAdminWeekObject'].replace('{Email}', Email)}",
			f"{Dictionary['Mail']['SucessAdminWeekBody'].replace('{Email}', Email).replace('{Customer.id}', Customer.id).replace('{VoucherID}', VoucherID).replace('{VoucherPassword}', VoucherPassword).replace('{ServerURL}', ServerURL).replace('{Location}', Location).replace('{Admin}', Admin)}"
		)
		ENVOYERMAILclient(
			Email,
			f"{Dictionary['Mail']['SucessClientWeekObject']}",
			f"{Dictionary['Mail']['SucessClientWeekBody'].replace('{VoucherID}', VoucherID).replace('{VoucherPassword}', VoucherPassword)}"
		)
	elif WeekOrMonth == "Month":
		print(Result.subscription)
		CodeDeleteplan = codecs.encode(Result.subscription, 'rot_13')  #This is to obfuscate the sub_xxx
		ENVOYERMAILAdmin(
			f"{Dictionary['Mail']['SucessAdminMonthObject'].replace('{Email}', Email)}",
			f"{Dictionary['Mail']['SucessAdminMonthBody'].replace('{Email}', Email).replace('{Customer.id}', Customer.id).replace('{VoucherID}', VoucherID).replace('{VoucherPassword}', VoucherPassword).replace('{ServerURL}', ServerURL).replace('{Location}', Location).replace('{CodeDeleteplan}', CodeDeleteplan)}"
		)
		ENVOYERMAILclient(
			Email,
			f"{Dictionary['Mail']['SucessClientMonthObject']}",
			f"{Dictionary['Mail']['SucessClientMonthBody'].replace('{VoucherID}', VoucherID).replace('{VoucherPassword}', VoucherPassword).replace('{ServerURL}', ServerURL).replace('{CodeDeleteplan}', CodeDeleteplan).replace('{Admin}', Admin)}"
		)

	#Write in PaymentHistory.xlsx
	heure = datetime.now()
	rowVide = getrowvidePaymentHistory()
	if WeekOrMonth == "Week":
		writeincellPaymentHistory("c", rowVide, "WEEK")  #Week ou month
		writeincellPaymentHistory("e", rowVide,Result.payment_intent)  #payment_intent
	if WeekOrMonth == "Month":
		writeincellPaymentHistory("c", rowVide, "MONTH")  #Week ou month
		writeincellPaymentHistory("e", rowVide, Result.subscription)  # sub_xxx
	writeincellPaymentHistory("b", rowVide, Location)  #Location
	writeincellPaymentHistory("d", rowVide, Email)  #Email
	writeincellPaymentHistory("f", rowVide, Customer.id)  #cus_xxx
	writeincellPaymentHistory("g", rowVide, VoucherID)  #VoucherID
	writeincellPaymentHistory("h", rowVide, VoucherPassword)  #VoucherPassword
	writeincellPaymentHistory("i", rowVide,f"{heure.hour}:{heure.minute} {heure.day}/{heure.month}/{heure.year}")  #Time he bought the plan
	PaymentHistoryload.save(PaymentHistoryXlsx)  #Save the modification
	print(f"{Dictionary['Print']['SavedInPaymentHistory']}")

	#Finally return html with info based on month or week
	if WeekOrMonth == "Week":
		return render_template(
			'Success.html',
			VoucherID=VoucherID,
			VoucherPassword=VoucherPassword,
			info=f"{Dictionary['Flask']['OneDeviceAtATime']}"
		)
	elif WeekOrMonth == "Month":
		return render_template(
			'Success.html',
			VoucherID=VoucherID,
			VoucherPassword=VoucherPassword,
			info=f"{Dictionary['Flask']['OneDeviceAtATime']}",
			info2=f"{Dictionary['Flask']['IfYouWantToDeleteYourPlanGoInYourMail']}"
		)


@app.route('/supprimermonacces')
def Delete():
	Location = request.headers['HTTP_X_LOCATION']
	Host = request.headers['HTTP_X_HOST']  #This is not really needed
	IP = request.headers['HTTP_X_REAL_IP']
	printRequest(IP, Host, str(request))
	sub_xxx = str(request.args.get('bangbang'))

	#try to decode using rot13 but it can fail if the user enter some chinese character for exemple idk
	try:
		sub_xxx = codecs.decode(sub_xxx, 'rot_13')
	except Exception as e:
		print(f"{Dictionary['Print']['ErrorRot_13'].replace('{sub_xxx}', sub_xxx).replace('{e}', str(e))}")
		return f"{Dictionary['Flask']['ErrorRot_13']}"

	#Basic check if it simply entered something else
	if not sub_xxx[0:4] == "sub_":
		print(f"{Dictionary['Print']['WrongSub_xxxEntered'].replace('{sub_xxx}', sub_xxx)}")
		return f"{Dictionary['Flask']['WrongSub_xxxEntered']}"

	#Now we really delete his plan
	try:
		sub = stripe.Subscription.retrieve(sub_xxx)
		sub.delete()
	except Exception as e:
		#If Stripe don't recognize the sub_xxx
		if "No such subscription" in str(e):
			print(f"{Dictionary['Print']['StripeDidntRecognizeThisSub_xxx'].replace('{sub_xxx}', sub_xxx).replace('{e}', str(e))}")
			return f"{Dictionary['Flask']['StripeDidntRecognizeThisSub_xxx']}"

		#If Stripe return something else
		print(f"{Dictionary['Print']['StripeCantDeleteThisSub_xxx'].replace('{sub_xxx}', sub_xxx).replace('{e}', str(e))}")
		ENVOYERMAILAdmin(
			f"{Dictionary['Mail']['StripeCantDeleteThisSub_xxxObject']}",
			f"{Dictionary['Mail']['StripeCantDeleteThisSub_xxxBody'].replace('{sub_xxx}', sub_xxx).replace('{e}', str(e))}"
		)
		return f"{Dictionary['Flask']['ErrorMailSentHereIsFicelle']}"

	try:
		nombrederow = getrowvidePaymentHistory()
		nombrederow = int(nombrederow) - 1#convert nombrederow to int to substract 1
		nombrederow = str(nombrederow)#Get back to normal
		jaitrouve = returncellcontentPaymentHistory("e",nombrederow)  # e pour sub_xxx
		if jaitrouve == sub_xxx:
			writeincellPaymentHistory("a", nombrederow, "DELETED")
			PaymentHistoryload.save(PaymentHistoryXlsx)  #savetout
	except Exception as e:
		print(f"{Dictionary['Print']['CantWriteDELETEDInPaymentHistory'].replace('{e}', str(e))}")
		ENVOYERMAILAdmin(
			f"{Dictionary['Mail']['CantWriteDELETEDInPaymentHistoryObject']}",
			f"{Dictionary['Mail']['CantWriteDELETEDInPaymentHistoryBody'].replace('{sub_xxx}', sub_xxx).replace('{e}', str(e))}"
		)

	#This is to prevent the user to re-use the same Stripe Session to look like he bought again
	with open("SessionID.json") as f:
		SessionIDJson = json.load(f)
		for cs_live in SessionIDJson.keys():
			if (SessionIDJson[cs_live]["Plan"] == "Month") and (SessionIDJson[cs_live]["sub_xxx"] == sub_xxx):
				LocationInSessionIDJson = SessionIDJson[cs_live]["Location"]
				SessionIDJson[cs_live]["Plan"] = "DELETED"
				break
		SessionIDJson = json.dumps(SessionIDJson,
								   indent=4,
								   sort_keys=True,
								   ensure_ascii=False)
		with open("SessionID.json", "w") as json_file:
			json_file.write(SessionIDJson)

	#And we send the mail
	InfoStripe = stripe.Subscription.retrieve(sub_xxx)
	Customer = stripe.Customer.retrieve(InfoStripe.customer)
	Email = Customer.email
	print(f"Bye bye {Email}")
	ENVOYERMAILAdmin(
		f"{Dictionary['Mail']['SuccessDeleteCustomerObject']}",
		f"{Dictionary['Mail']['SuccessDeleteCustomerBody'].replace('{Email}', Email).replace('{LocationInSessionIDJson}', LocationInSessionIDJson).replace('{Location}', Location).replace('{sub_xxx}', sub_xxx).replace('{InfoStripe.customer}', InfoStripe.customer).replace('{InfoStripe}', InfoStripe).replace('{Customer}', Customer)}"
	)
	return f"{Dictionary['Flask']['SuccessDeleteCustomer']}"


@app.route('/logo')
def logo():
	return send_file("static/image/logo.png")


@app.route('/background')
def background():
	return send_file("static/image/background.webp")


if __name__ == "__main__":
	# app.jinja_env.auto_reload = True
	# app.config['TEMPLATES_AUTO_RELOAD'] = True
	app.run(host='127.0.0.1', port=5000)
